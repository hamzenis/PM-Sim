"""
Fast Simulation Script
This script runs different types of simulations (random, fixed, and fixed continuous)
and saves the results as CSV files.

For Debug: It also combines the CSV files into a single Excel sheet
"""

from statistics import mean
from io import StringIO
from pathlib import Path

from softDsim.settings import DEBUG_DIRECTORY
import pandas as pd
import openpyxl

from fast_simulation_framework.fastsim_helpers import (
    init_scenario,
    init_config,
    init_skill_types,
    init_members,
    set_scenario_state,
    set_members,
    set_tasks_fixed,
    set_tasks_random,
    run_simulation,
    run_continuous_simulation,
    NpRecord,
)

# User parameters(Workpacks) for simulations
import fast_simulation_framework.workpacks as wp

# Keep imports used by referencing them in a mapping
AVAILABLE_WORKPACKS = {
    "only_tasks": wp.USERPARAMETERS_ONLY_TASKS,
    "example": wp.USERPARAMETERS_EXAMPLE,
    "original": wp.USERPARAMETERS_OG,
}

# Configuration for simulation runs
RUNS_OF_SIM = 1
CREATE_EVERY_RUN_CSV = 1
OUTPUT_DIR = DEBUG_DIRECTORY
UPs = AVAILABLE_WORKPACKS["only_tasks"]

# Constant for holding CSV file names, for the excel table
CSV_LIST = [
    "capacity_calc_stats.csv",
    "fastsim_softdsim.csv",
    "task_work_log.csv",
    "familiarity_calc_stats.csv",
]


def fixed_fastsimulation_with_continuous_simulation():
    """
    Fixed fast simulation for generating consistent simulation data.
    It uses fixed parameters for scenario configuration, skill types, and tasks.
    It continues the simulation with each workpack instead of resetting the scenario.
    """
    print("Started fixed fastsimulation with continuous simulation")
    is_simulation_fixed = True
    is_first_run = True
    rec = NpRecord()
    scenario, state, team = init_scenario()
    config = init_config(is_simulation_fixed)
    skill_types = init_skill_types(is_simulation_fixed)
    members = init_members(skill_types)

    for x in range(1, RUNS_OF_SIM + 1):
        for n, UP in enumerate(UPs):
            if is_first_run:
                state = set_scenario_state(state)
                members = set_members(members)
                tasks = set_tasks_fixed(scenario)
                scenario = run_continuous_simulation(
                    scenario, config, members, tasks, skill_types, rec, UP, n
                )
                is_first_run = False
            else:
                scenario = run_continuous_simulation(
                    scenario, config, members, tasks, skill_types, rec, UP, n
                )

            # print(f"{len(tasks.done())} \t {mean([m.efficiency for m in members])}")

        if x % CREATE_EVERY_RUN_CSV == 0:
            print(f"{x} of {RUNS_OF_SIM}")
            csv_buffer = StringIO()
            rec.df().to_csv(csv_buffer)

            try:
                OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            except Exception as e:
                print(f"Failed to create output directory {OUTPUT_DIR}: {e}")

            local_filename = OUTPUT_DIR / "fastsim_softdsim.csv"

            try:
                with open(local_filename, "w") as f:
                    f.write(csv_buffer.getvalue())
                print(f"Wrote CSV to {local_filename}")
            except Exception as e:
                print(f"Failed to write CSV to {local_filename}: {e}")

            rec.clear()


def fixed_fastsimulation():
    """
    Fixed fast simulation for generating consistent simulation data.
    It uses fixed parameters for scenario configuration, skill types, and tasks.
    It executes independent simulations for each workpack.
    """
    print("Started fixed fastsimulation")
    is_simulation_fixed = True
    rec = NpRecord()
    scenario, state, team = init_scenario()
    config = init_config(is_simulation_fixed)
    skill_types = init_skill_types(is_simulation_fixed)
    members = init_members(skill_types)
    for x in range(1, RUNS_OF_SIM + 1):
        for n, UP in enumerate(UPs):
            state = set_scenario_state(state)
            members = set_members(members)
            tasks = set_tasks_fixed(scenario)
            run_simulation(scenario, config, members, tasks, skill_types, rec, UP, n)
            print(f"{len(tasks.done())} \t {mean([m.efficiency for m in members])}")

        if x % CREATE_EVERY_RUN_CSV == 0:
            print(f"{x} of {RUNS_OF_SIM}")
            csv_buffer = StringIO()
            rec.df().to_csv(csv_buffer)

            try:
                OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            except Exception as e:
                print(f"Failed to create output directory {OUTPUT_DIR}: {e}")

            local_filename = OUTPUT_DIR / "fastsim_softdsim.csv"

            try:
                with open(local_filename, "w") as f:
                    f.write(csv_buffer.getvalue())
                print(f"Wrote CSV to {local_filename}")
            except Exception as e:
                print(f"Failed to write CSV to {local_filename}: {e}")

            rec.clear()


def random_fastsimulation():
    """
    Random fast simulation for generating diverse simulation data.
    It uses random parameters for scenario configuration, skill types, and tasks.
    It executes independent simulations for each workpack.
    """
    print("Started random fastsimulation")
    is_simulation_fixed = False
    rec = NpRecord()
    scenario, state, team = init_scenario()
    config = init_config()
    skill_types = init_skill_types(is_simulation_fixed)
    members = init_members(skill_types)
    for x in range(1, RUNS_OF_SIM + 1):
        for n, UP in enumerate(UPs):
            state = set_scenario_state(state)
            members = set_members(members)
            tasks = set_tasks_random(scenario)
            run_simulation(scenario, config, members, tasks, skill_types, rec, UP, n)
            print(f"{len(tasks.done())} \t {mean([m.efficiency for m in members])}")

        if x % CREATE_EVERY_RUN_CSV == 0:
            print(f"{x} of {RUNS_OF_SIM}")
            csv_buffer = StringIO()
            rec.df().to_csv(csv_buffer)

            try:
                OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            except Exception as e:
                print(f"Failed to create output directory {OUTPUT_DIR}: {e}")

            local_filename = OUTPUT_DIR / "fastsim_softdsim.csv"

            try:
                with open(local_filename, "w") as f:
                    f.write(csv_buffer.getvalue())
                print(f"Wrote CSV to {local_filename}")
            except Exception as e:
                print(f"Failed to write CSV to {local_filename}: {e}")

            rec.clear()


def csvs_to_excel_sheet():
    """
    Combine all CSV files in the output directory into a single Excel sheet.
    All CSV Files are imported side by side with one blank column between them.
    """
    output_dir = Path(OUTPUT_DIR)
    if not output_dir.exists():
        print(f"No output directory found at {output_dir}, skipping Excel merge.")
        return
    csv_paths = []
    for name in CSV_LIST:
        path = output_dir / name
        if path.is_file():
            csv_paths.append(path)
        else:
            print(f"CSV file '{name}' not found in {output_dir}, skipping.")
    if not csv_paths:
        print(f"No CSV files found in {output_dir}, skipping Excel merge.")
        return
    excel_filename = output_dir / "combined_simulation_data.xlsx"
    try:
        with pd.ExcelWriter(excel_filename, engine="xlsxwriter") as writer:
            sheet_name = "SimulationData"
            startcol = 0
            for csv_path in csv_paths:
                try:
                    df = pd.read_csv(csv_path)
                except Exception as e:
                    print(f"Failed to read {csv_path}: {e}")
                    continue
                # write each CSV into the same sheet at increasing startcol (side by side)
                df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    startrow=0,
                    startcol=startcol,
                    index=False,
                )
                startcol += df.shape[1] + 1
        print(f"Combined Excel file created at {excel_filename}")
        return
    except Exception as e:
        print(f"Failed to write Excel with XlsxWriter: {e}")


def insert_score_formulas():
    """
    Insert score calculation formulas into the Excel sheet.
    Based on a 12 week workpack simulation structure and a budget of 250000
    TODO: Adjust to use dynamic cell references, like workpack length
    """
    quality_score_formula = "=IF(AU13=0, 0, ABS((1 - (AT13/AU13))^1.0 * 100))"
    budget_score_formula = "=IF(AL13<=250000,100,MAX(0, (100 - (( (AL13/250000) - 1 ) * 100 )^1)) * 100/100)"
    time_score_formula = (
        "=IF(AM13<=60,100,MAX(0, (100 - (( (AM13/60) - 1 ) * 100 )^1)) * 100 / 100)"
    )
    total_score_formula = "=ROUND((K19 + L19 + M19) / (300) * 100,2)"

    excel_filename = OUTPUT_DIR / "combined_simulation_data.xlsx"

    # Ensure file exists
    if not excel_filename.exists():
        print(
            f"No combined Excel file found at {excel_filename}, skipping score insertion."
        )
        return

    try:
        # Load workbook and pick the SimulationData sheet if present
        wb = openpyxl.load_workbook(excel_filename)
        sheet_name = "SimulationData"
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb[wb.sheetnames[0]]

        # Write headers for the three scores (row 18)
        ws["K18"] = "Quality Score"
        ws["L18"] = "Budget Score"
        ws["M18"] = "Time Score"
        ws["N18"] = "Total Score"

        # Insert formulas into row 19 for each score
        ws["K19"] = quality_score_formula
        ws["L19"] = budget_score_formula
        ws["M19"] = time_score_formula
        ws["N19"] = total_score_formula

        # Save workbook
        wb.save(excel_filename)
        print(f"Inserted score formulas into {excel_filename}")
    except Exception as e:
        print(f"Failed to insert score formulas into {excel_filename}: {e}")


def run():
    """
    Function to run different fast simulations.

    How to run this script:
    1. Navigate to the backend directory of the project.
    2. Activate the virtual environment
    3. Run the script using Django's runscript command:
        python manage.py runscript fastsim
    """

    try:
        prompt = (
            "Select fast simulation mode:\n"
            "  1 - fixed continuous (default)\n"
            "  2 - fixed\n"
            "  3 - random\n"
            "Enter choice [1]: "
        )
        choice = input(prompt).strip().lower()
    except (EOFError, KeyboardInterrupt):
        print("No input detected, defaulting to '1' (fixed continuous).")
        choice = "1"

    if choice == "1":
        print("Running: fixed_fastsimulation_with_continuous_simulation")
        fixed_fastsimulation_with_continuous_simulation()
    elif choice == "2":
        print("Running: fixed_fastsimulation")
        fixed_fastsimulation()
    elif choice == "3":
        print("Running: random_fastsimulation")
        random_fastsimulation()
    else:
        print(f"Unrecognized choice '{choice}', defaulting to fixed continuous.")
        fixed_fastsimulation_with_continuous_simulation()

    csvs_to_excel_sheet()
    insert_score_formulas()
